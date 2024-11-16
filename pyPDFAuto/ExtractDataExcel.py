import openpyxl
import ProcesingData as PS
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import DataValidation as DataV


class workbookPath:
    Excelpath = ""
    Publipath = "E:/Publicadores/"
    Auxpath = "E:/Publicadores/PRECURSORES AUXILIARES/"
    Regpath = "E:/Publicadores/PRECURSORES REGULARES/"
    def inputExcelpath(file):
        workbookPath.Excelpath = file
    def inputPublipath(file):
        workbookPath.Publipath = file + '/'
    def inputAuxpath(file):
        workbookPath.Auxpath = file + '/'
    def inputRegpath(file):
        workbookPath.Regpath = file + '/'        
def openFile():
    ventana2 = tk.Tk()
    ventana2.geometry("350x200")
    ventana2.resizable(False, False)
    ventana2.configure(background="white")
    btnpublicador = tk.Button(ventana2, text="Publicador", command=rootPubli)
    btnpublicador.grid(row = 1, column = 0)
    btnprecursorAux = tk.Button(ventana2, text="Precursor auxiliar", command= rootAux)
    btnprecursorAux.grid(row = 1, column = 1)
    btnprecursorReg = tk.Button(ventana2, text="Precursor regular", command=rootReg)
    btnprecursorReg.grid(row = 1, column = 2)
    btnExcel= tk.Button(ventana2, text="Excel", command=rootExcel)
    btnExcel.grid(row = 1, column = 3)
    ventana2.mainloop()

def rootExcel():
    try:
        filepath = filedialog.askopenfilename(initialdir="E:/Descargas/", title="Abrir Archivo Excel?", filetypes=(('Excel', '.xlsx'), ('MacroExcel', '.xlsm'), ('BinarioExcel', '.xlsb'), ('OpenDocument', '.ods')))
        print(messagebox.showinfo(message='Archivo seleccionado: '+ filepath, title='info'))
        workbookPath.inputExcelpath(filepath)           
    except: 
        FileNotFoundError(print(messagebox.showinfo(message='No selecciono archivo', title='Error')))      
def rootPubli():
    try:
        filepath = filedialog.askdirectory(initialdir="E:/Publicadores/", title="Abrir ruta publicadores?")
        print(messagebox.showinfo(message='Ruta seleccionada: '+ filepath, title='info'))
        workbookPath.inputPublipath(filepath)           
    except: 
        FileNotFoundError(print(messagebox.showinfo(message='No selecciono archivo', title='Error')))
def rootAux():
    try:
        filepath = filedialog.askdirectory(initialdir="E:/Publicadores/PRECURSORES AUXILIARES/", title="Abrir ruta Auxiliares?")
        print(messagebox.showinfo(message='Ruta seleccionada: '+ filepath, title='info'))
        workbookPath.inputAuxpath(filepath)           
    except: 
        FileNotFoundError(print(messagebox.showinfo(message='No selecciono archivo', title='Error')))
def rootReg():
    try:
        filepath = filedialog.askdirectory(initialdir="E:/Publicadores/PRECURSORES REGULARES/", title="Abrir ruta Regukares?")
        print(messagebox.showinfo(message='Ruta seleccionada: '+ filepath, title='info'))
        workbookPath.inputRegpath(filepath)           
    except: 
        FileNotFoundError(print(messagebox.showinfo(message='No selecciono archivo', title='Error')))

def ExtractInfo():
    wb = openpyxl.load_workbook(str(workbookPath.Excelpath))
    ws = wb.active
    columDataNames = txtname.get()
    columDataNames = columDataNames.upper()
    columDataPreach = txtpreach.get()
    columDataPreach = columDataPreach.upper()
    columDataCourses = txtcourses.get()
    columDataCourses =  columDataCourses.upper()
    columDataNotes = txtnotes.get()
    columDataNotes = columDataNotes.upper()
    starDataNames = int(txtstar.get())
    endDataNames = int(txtend.get())
    month = txtmonth.get() 
    month = month.lower()
    DataFull = DataV.PDF_DataValidationFull(columDataNames, month, columDataPreach, columDataCourses, columDataNotes, starDataNames, endDataNames)
    DataNormal = DataV.PDF_DataValidationNormal(columDataNames, month, columDataPreach, columDataCourses, columDataNotes, starDataNames, endDataNames)

    if int(var.get()) == 1:
        if DataNormal == False:
          
            print(messagebox.showerror(message=str(DataNormal[1]), title='Error'))

        elif DataNormal == True: 

            while starDataNames <= endDataNames:
                name = ws[columDataNames + str(starDataNames)].value
                preach = ws[columDataPreach + str(starDataNames)].value
                courses = ws[columDataCourses + str(starDataNames)].value
                notes = ws[columDataNotes + str(starDataNames)].value
                name = str(name.upper())
                root = str(workbookPath.Publipath)
                PS.PDF_RowSetNormal(root,name, month, preach, courses, notes)
                starDataNames = starDataNames + 1

    elif int(var.get()) == 2:
        if DataFull == False:
          
            print(messagebox.showerror(message=str(DataFull[1]), title='Error'))

        elif DataFull == True: 
      
            while starDataNames <= endDataNames:
                name = ws[columDataNames + str(starDataNames)].value
                courses = ws[columDataCourses + str(starDataNames)].value
                notes = ws[columDataNotes + str(starDataNames)].value
                hours = ws[columDataPreach + str(starDataNames)].value
                name = str(name.upper())
                root = str(workbookPath.Auxpath)
                PS.PDF_RowSetAux(root, name, month, courses, hours, notes)
                starDataNames = starDataNames + 1

    elif int(var.get()) == 3:
        if DataFull == False:
          
            print(messagebox.showerror(message=str(DataFull[1]), title='Error'))

        elif DataFull == True: 
      
            while starDataNames <= endDataNames:
                name = ws[columDataNames + str(starDataNames)].value
                courses = ws[columDataCourses + str(starDataNames)].value
                notes = ws[columDataNotes + str(starDataNames)].value
                hours = ws[columDataPreach + str(starDataNames)].value
                name = str(name.upper())
                root = str(workbookPath.Regpath)
                PS.PDF_RowSetColporter(root, name, month, courses, hours, notes)
                starDataNames = starDataNames + 1


ventana = tk.Tk()
ventana.geometry("480x300")
ventana.resizable(False, False)
ventana.configure(background="white")
ventana.title("Auto PDF tarjetas predicacion")
tk.Wm.wm_title(ventana, "Auto preach filler")

lblname = tk.Label(ventana, text="Columna Nombre").grid(row = 0, column = 0)
txtname = tk.Entry(ventana)
txtname.grid(row = 0, column = 1)
lblpreach = tk.Label(ventana, text="Columna Predicacion").grid(row = 0, column = 2)
txtpreach = tk.Entry(ventana)
txtpreach.grid(row = 0, column = 3)
lblcourses = tk.Label(ventana, text="Columna Cursos").grid(row = 1, column = 0)
txtcourses = tk.Entry(ventana)
txtcourses.grid(row = 1, column = 1)
lblnotes = tk.Label(ventana, text="Columna Notas").grid(row = 1, column = 2)
txtnotes = tk.Entry(ventana)
txtnotes.grid(row = 1, column = 3)
lblstar = tk.Label(ventana, text="Celda inicial").grid(row = 2, column = 0)
txtstar = tk.Entry(ventana)
txtstar.grid(row = 2, column = 1)
lblend = tk.Label(ventana, text="Celda final").grid(row = 2, column = 2)
txtend = tk.Entry(ventana)
txtend.grid(row = 2, column = 3)
lblmonth = tk.Label(ventana, text="Mes").grid(row = 3, column = 0)
txtmonth = tk.Entry(ventana)
txtmonth.grid(row = 3, column = 1)
btniniciar = tk.Button(ventana, text="Iniciar grabado", command=ExtractInfo)
btniniciar.grid(row = 5, column = 1)
photoImage = tk.PhotoImage(file = r"C:\Users\joqta\OneDrive\Documentos\GitHub\PDFAutoFillCop\pyPDFAuto\Iconfile.png")
photo = photoImage.subsample(3, 3)
btnBuscar = tk.Button(ventana, image=photo, command=openFile)
btnBuscar.grid(row = 4, column = 3)

var = tk.IntVar()
R1 = tk.Radiobutton(ventana, text="Publicador", variable=var, value=1)
R1.grid(row = 4, column = 0)
R2 = tk.Radiobutton(ventana, text="Precursor auxiliar", variable=var, value=2)
R2.grid(row = 4, column = 1)
R3 = tk.Radiobutton(ventana, text="Precursor regular", variable=var, value=3)
R3.grid(row = 4, column = 2)

ventana.mainloop()
