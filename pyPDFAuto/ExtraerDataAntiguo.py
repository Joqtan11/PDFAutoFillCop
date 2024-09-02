import openpyxl
import ProcesingData as PS

wb = openpyxl.load_workbook('E:/Descargas/julio 2024.xlsx')
ws = wb.active

print ('Ingrese la columna a trabajar en nombres')
columDataNames = input()
print ('Ingrese la columna a trabajar de verificacion de predicacion')
columDataPreach = input()
print ('Ingrese la columna a trabajar de cursos hechos')
columDataCourses = input()
print ('Ingrese la columna a trabajar de notas')
columDataNotes = input()
print ('Ingrese el numero de celda inicial')
starDataNames = int(input())
print ('Ingrese el numero de celda final')
endDataNames = int(input())
print ('Ingrese el mes')
month = input()

while starDataNames <= endDataNames:
   name = ws[columDataNames + str(starDataNames)].value
   preach = ws[columDataPreach + str(starDataNames)].value
   courses = ws[columDataCourses + str(starDataNames)].value
   notes = ws[columDataNotes + str(starDataNames)].value + ' '
   name = str(name.upper())
   PS.PDF_RowSetNormal(name, month, preach, courses, notes)
   starDataNames = starDataNames + 1 

print ('Ingrese la columna a trabajar en nombres de auxiliares')
columDataNames = input()
print ('Ingrese la columna a trabajar de horas')
columDatahours = input()
print ('Ingrese la columna a trabajar de cursos hechos')
columDataCourses = input()
print ('Ingrese la columna a trabajar de notas')
columDataNotes = input()
print ('Ingrese el numero de celda inicial')
starDataNames = int(input())
print ('Ingrese el numero de celda final')
endDataNames = int(input())
print ('Ingrese el mes')
month = input()

while starDataNames <= endDataNames:
   name = ws[columDataNames + str(starDataNames)].value
   courses = ws[columDataCourses + str(starDataNames)].value
   notes = ws[columDataNotes + str(starDataNames)].value + ' '
   hours = ws[columDatahours+ str(starDataNames)].value
   name = str(name.upper())
   PS.PDF_RowSetAux(name, month, courses, hours, notes)
   starDataNames = starDataNames + 1

print ('Ingrese la columna a trabajar en nombres de precursores')
columDataNames = input()
print ('Ingrese la columna a trabajar de horas')
columDatahours = input()
print ('Ingrese la columna a trabajar de cursos hechos')
columDataCourses = input()
print ('Ingrese la columna a trabajar de notas')
columDataNotes = input()
print ('Ingrese el numero de celda inicial')
starDataNames = int(input())
print ('Ingrese el numero de celda final')
endDataNames = int(input())
print ('Ingrese el mes')
month = input()

while starDataNames <= endDataNames:
   name = ws[columDataNames + str(starDataNames)].value
   courses = ws[columDataCourses + str(starDataNames)].value
   notes = ws[columDataNotes + str(starDataNames)].value + ' '
   hours = ws[columDatahours+ str(starDataNames)].value
   name = str(name.upper())
   PS.PDF_RowSetColporter(name, month, courses, hours, notes)
   starDataNames = starDataNames + 1